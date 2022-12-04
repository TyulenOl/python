import csv
import math
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from datetime import datetime
import doctest

dic_naming = {'name': 'Название', 'description': 'Описание', 'key_skills': 'Навыки', 'experience_id': 'Опыт работы',
              'premium': 'Премиум-вакансия', 'employer_name': 'Компания', 'salary_from': 'Оклад',
              'area_name': 'Название региона', 'published_at': 'Дата публикации вакансии'}
currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055}
substitution_work_experience = {"noExperience": "Нет опыта", "between1And3": "От 1 года до 3 лет",
                                "between3And6": "От 3 до 6 лет", "moreThan6": "Более 6 лет"}
substitution_currency = {"AZN": "Манаты", "BYR": "Белорусские рубли", "EUR": "Евро", "GEL": "Грузинский лари",
                         "KGS": "Киргизский сом", "KZT": "Тенге", "RUR": "Рубли", "UAH": "Гривны", "USD": "Доллары",
                         "UZS": "Узбекский сум"}
reverse_naming = {'Навыки': 'key_skills', 'Оклад': 'salary', 'Дата публикации вакансии': 'published_at',
                  'Опыт работы': 'experience_id', 'Премиум-вакансия': 'premium',
                  'Идентификатор валюты оклада': 'salary_currency', 'Название': 'name',
                  'Название региона': 'area_name', 'Компания': 'employer_name', 'Описание': 'description'}


class DataSet:
    """Класс для хранения и обработки данных и статистики.

    Attributes:
        file_name (str): Имя исходного файла с данными
        vacancies_objects (list[Vacancy]): Лист вакансий со всеми заполненными значениями
        statistic (list[dict[int: int or str: int]]): Валюта оклада
    """

    def __init__(self, file_name, vacancies_objects):
        """Инициализирует объект DataSet.

        Args:
            file_name (str): Имя исходного файла с данными
            vacancies_objects (list): Лист вакансий для обработки

        >>> type(DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']])).__name__
        'DataSet'
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).file_name
        'unittest.csv'
        >>> len(DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects)
        1
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].name
        'IT аналитик'
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].area_name
        'Санкт-Петербург'
        >>> type(DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].published_at).__name__
        'datetime'
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].published_at.strftime("%Y-%m-%dT%H:%M:%S%z")
        '2007-12-03T17:34:36+0300'
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].published_at.strftime("%Y.%m.%d")
        '2007.12.03'
        >>> type(DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].salary).__name__
        'Salary'
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].salary.salary_from
        35000
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].salary.salary_to
        45000
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).vacancies_objects[0].salary.salary_currency
        'RUR'
        """

        self.file_name = file_name
        self.vacancies_objects = [Vacancy(row) for row in vacancies_objects if None not in row and '' not in row]
        self.statistic = []

    def calculate_statistics(self, profession_name):
        """Вычисляет статистику по вакансиям: динамика уровня зарплат по годам, динамика количества вакансий по
        годам, динамика уровня зарплат по годам для выбранной профессии, динамика количества вакансий по годам для
        выбранной профессии, уровень зарплат по городам (в порядке убывания) - только первые 10 значений,
        доля вакансий по городам (в порядке убывания) - только первые 10 значений.

        Args:
            profession_name(str): Название професии для сбора более конкретной статистики по данной професии

        Returns:
            list[dict[int: int or str: int]]: Собранная сатистика: динамика уровня зарплат по годам, динамика количества
             вакансий по годам, динамика уровня зарплат по годам для выбранной профессии, динамика количества вакансий
             по годам для выбранной профессии, уровень зарплат по городам (в порядке убывания) - только первые 10
             значений, доля вакансий по городам (в порядке убывания) - только первые 10 значений

        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).calculate_statistics('аналитик')
        [{2007: 40000, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {2007: 1, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {2007: 40000, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {2007: 1, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {'Санкт-Петербург': 40000}, {'Санкт-Петербург': 1.0}]
        >>> DataSet('unittest.csv', [['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']]).calculate_statistics('Инженер')
        [{2007: 40000, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {2007: 1, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {2007: 0, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {2007: 0, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}, {'Санкт-Петербург': 40000}, {'Санкт-Петербург': 1.0}]
        """

        salary_by_years = {}
        salary_by_years_profession = {}
        sum_salary_by_city = {}
        salary_by_city = {}
        number_vac_by_years = {}
        number_profession_by_years = {}
        number_vac_by_city = {}
        percentage_vac_by_city = {}
        for vacancy in self.vacancies_objects:
            year = vacancy.published_at.year
            city = vacancy.area_name
            number_vac_by_years[year] = number_vac_by_years.get(year, 0) + 1
            salary_by_years[year] = salary_by_years.get(year, 0) + vacancy.salary.convert_to_rubles()
            number_vac_by_city[city] = number_vac_by_city.get(city, 0) + 1
            sum_salary_by_city[city] = sum_salary_by_city.get(city, 0) + vacancy.salary.convert_to_rubles()
            salary_by_years_profession.setdefault(year, 0)
            number_profession_by_years.setdefault(year, 0)
            if profession_name in vacancy.name:
                salary_by_years_profession[year] = salary_by_years_profession.get(year,
                                                                                  0) + vacancy.salary.convert_to_rubles()
                number_profession_by_years[year] = number_profession_by_years.get(year, 0) + 1

        for year in range(2007, 2023):
            if year in number_vac_by_years.keys():
                salary_by_years[year] = math.floor(salary_by_years[year] / number_vac_by_years[year])
            else:
                number_vac_by_years[year] = 0
                salary_by_years[year] = 0

        for year in range(2007, 2023):
            if year in number_profession_by_years.keys():
                if number_profession_by_years[year] != 0:
                    salary_by_years_profession[year] = math.floor(
                        salary_by_years_profession[year] / number_profession_by_years[year])
            else:
                number_profession_by_years[year] = 0
                salary_by_years_profession[year] = 0

        for city in number_vac_by_city.keys():
            proportion_vacancy = number_vac_by_city.get(city) / len(self.vacancies_objects)
            if proportion_vacancy >= 0.01:
                percentage_vac_by_city[city] = round(proportion_vacancy, 4)
                salary_by_city[city] = math.floor(sum_salary_by_city[city] / number_vac_by_city.get(city))

        sorted_salary_by_years = dict(sorted(salary_by_years.items(), key=lambda x: x[0]))
        sorted_number_vac_by_years = dict(sorted(number_vac_by_years.items(), key=lambda x: x[0]))
        sorted_salary_by_years_profession = dict(sorted(salary_by_years_profession.items(), key=lambda x: x[0]))
        sorted_number_profession_by_years = dict(sorted(number_profession_by_years.items(), key=lambda x: x[0]))
        sorted_salary_by_city = dict(sorted(salary_by_city.items(), key=lambda x: -x[1])[:10])
        sorted_percentage_vac_by_city = dict(sorted(percentage_vac_by_city.items(), key=lambda x: -x[1])[:10])

        self.statistic = [sorted_salary_by_years, sorted_number_vac_by_years, sorted_salary_by_years_profession,
                          sorted_number_profession_by_years, sorted_salary_by_city, sorted_percentage_vac_by_city]

        return self.statistic

    def print_statistic(self):
        """Печатает статистику: динамика уровня зарплат по годам,динамика количества вакансий по годам,
        динамика уровня зарплат по годам для выбранной профессии, динамика количества вакансий по годам для выбранной
        профессии, уровень зарплат по городам (в порядке убывания) - только первые 10 значений, доля вакансий по
        городам (в порядке убывания) - только первые 10 значений.
        """

        print(f'Динамика уровня зарплат по годам: {self.statistic[0]}')
        print(f'Динамика количества вакансий по годам: {self.statistic[1]}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.statistic[2]}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.statistic[3]}')
        print(f'Уровень зарплат по городам (в порядке убывания): {self.statistic[4]}')
        print(f'Доля вакансий по городам (в порядке убывания): {self.statistic[5]}')


class Vacancy:
    """Класс для представления вакансии.

    Attributes:
        name (str): Название профессии
        salary (Salary): Оклад
        area_name (): Название региона
        published_at (): Дата публикации вакансии
    """

    def __init__(self, vacancy):
        """Устанавливает все необходимые атрибуты для объекта Vacancy.

        Args: vacancy (list): Лист данных о вакансии состоящий из: название профессии, оклад, название региона, дата
        публикации вакансии.

        >>> type(Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300'])).__name__
        'Vacancy'
        >>> Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).name
        'IT аналитик'
        >>> Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).area_name
        'Санкт-Петербург'
        >>> type(Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).published_at).__name__
        'datetime'
        >>> Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).published_at.strftime("%Y-%m-%dT%H:%M:%S%z")
        '2007-12-03T17:34:36+0300'
        >>> type(Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).salary).__name__
        'Salary'
        >>> Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).salary.salary_from
        35000
        >>> Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).salary.salary_to
        45000
        >>> Vacancy(['IT аналитик', '35000.0', '45000.0','RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300']).salary.salary_currency
        'RUR'
        """

        self.name = vacancy[0].replace('\xa0', '\x20')
        self.salary = Salary([vacancy[1], vacancy[2], vacancy[3]])
        self.area_name = vacancy[4]
        self.published_at = datetime.strptime(vacancy[5], "%Y-%m-%dT%H:%M:%S%z")


class Salary:
    """Класс для представления оклада.

    Attributes:
        salary_from (int): Нижняя граница оклада
        salary_to (int): Верхняя граница оклада
        salary_currency (str): Валюта оклада
    """

    def __init__(self, salary):
        """Инициализирует объект Salary.

        Args:
            salary_from (str or int or float): Нижняя граница оклада
            salary_to (str or int or float): Нижняя граница оклада
            salary_currency (str): Нижняя граница оклада

        >>> type(Salary(['123000.0', '987000.0','RUR'])).__name__
        'Salary'
        >>> Salary(['123000.0', '987000.0','RUR']).salary_from
        123000
        >>> Salary(['123000', '987000.0','RUR']).salary_from
        123000
        >>> Salary(['123000.0', '987000.0','RUR']).salary_to
        987000
        >>> Salary(['123000.0', '987000','RUR']).salary_to
        987000
        >>> Salary(['123000.0', '987000.0','RUR']).salary_currency
        'RUR'
        >>> Salary(['123000.0', '987000.0','AZN']).salary_currency
        'AZN'
        """

        self.salary_from = int(float(salary[0]))
        self.salary_to = int(float(salary[1]))
        self.salary_currency = salary[2]

    def convert_to_rubles(self):
        """Вычисляет среднее значение зарплаты и конвертирует в рубли, при помощи словоря - currency_to_rub.

        Returns:
            float: Cреднее значение зарплаты в рублях

        >>> Salary(['123000.0', '987000.0','RUR']).convert_to_rubles()
        555000.0
        >>> Salary(['123000', '987000.0','RUR']).convert_to_rubles()
        555000.0
        >>> Salary(['123000.0', '987000','RUR']).convert_to_rubles()
        555000.0
        >>> Salary(['123000.0', '987000.0','AZN']).convert_to_rubles()
        19802400.0
        """

        return ((float(self.salary_from) + float(self.salary_to)) / 2) * float(currency_to_rub[self.salary_currency])


class Report:
    """Класс для формирования отчета в табличном виде.

    Attributes:
        sheet_titles (list[str]): Названия листов таблицы
        wb (Workbook): Рабочая книга
    """

    def __init__(self, sheet_titles, headlines):
        """Создает рабочую книгу и листы. У листов задает названия и заполняет первую строчку заголовков.

        Args:
            sheet_titles (list[str]): Названия листов таблицы
            headlines (list[dict[str: str]]): Заголовки, которые присваиваются определенным столбцам в первой строчке
        """

        self.sheet_titles = sheet_titles
        self.wb = Workbook()
        self.wb.remove(self.wb['Sheet'])
        for title in sheet_titles:
            self.wb.create_sheet(title)

        for sheet in range(len(headlines)):
            ws = self.wb[self.wb.sheetnames[sheet]]
            ws.append(headlines[sheet])

    def setting_workbook(self):
        """Настраивает рабочую книгу задавая каждому листу стили: называния столбцов выделены полужирным шрифтом,
        все ячейки с данными имеют тонкую границу черного цвета, ширина столбцов устанавливается для вместимости
        самой длинной строки в столбце, для столбца "Доля вакансий" устанавливается процентный формат данных."""
        for ws in self.wb:
            for column_cells in ws.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    ws.column_dimensions[new_column_letter].width = new_column_length * 1.20

            side = Side(style='thin', color="000000")
            for cell in ws._cells.values():
                if cell.row == 1:
                    cell.font = Font(bold=True, size=11)
                if cell.value is not None:
                    cell.border = Border(top=side, bottom=side, left=side, right=side)

        for cell in self.wb[self.wb.sheetnames[1]]['E']:
            cell.number_format = FORMAT_PERCENTAGE_00

    def generate_excel(self, statistic):
        """Создает в каталоге таблицу со статистикой.

        Args:
            statistic (list[dict[int: int or str: int]]): Статистика, которая должна выводиться в таблице
        """
        salary_by_years = statistic[0]
        number_vac_by_years = statistic[1]
        salary_by_years_profession = statistic[2]
        number_profession_by_years = statistic[3]
        salary_by_city = statistic[4]
        percentage_vac_by_city = statistic[5]
        ws = self.wb[self.wb.sheetnames[0]]
        row = 2
        for year in number_vac_by_years.keys():
            ws[row][0].value = year
            ws[row][1].value = salary_by_years[year]
            ws[row][2].value = salary_by_years_profession[year]
            ws[row][3].value = number_vac_by_years[year]
            ws[row][4].value = number_profession_by_years[year]
            row += 1

        ws = self.wb[self.wb.sheetnames[1]]
        row = 2
        for city in salary_by_city.keys():
            ws[row][0].value = city
            ws[row][1].value = salary_by_city[city]
            row += 1

        row = 2
        for city in percentage_vac_by_city.keys():
            ws[row][3].value = city
            ws[row][4].value = percentage_vac_by_city[city]
            row += 1

        self.setting_workbook()
        self.wb.save('report.xlsx')
        self.wb.close()


def csv_reader(file_name):
    """Открывает и читает необходимый файл, а также возвращяет полученный результат.

    Args:
       file_name (str): Название файла для чтения

    Returns:
        DataSet, list: Полученные данные из прочитанного файла, строчка с названиями столбцов

    >>> len(csv_reader('unittest.csv'))
    2
    >>> type(csv_reader('unittest.csv')[0]).__name__
    'DataSet'
    >>> type(csv_reader('unittest.csv')[1]).__name__
    'list'
    >>> csv_reader('unittest.csv')[1]
    ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']
    >>> csv_reader('unittest.csv')[0].file_name
    'unittest.csv'
    >>> len(csv_reader('unittest.csv')[0].vacancies_objects)
    3
    >>> type(csv_reader('unittest.csv')[0].vacancies_objects[0]).__name__
    'Vacancy'
    >>> csv_reader('unittest.csv')[0].vacancies_objects[0].name
    'IT аналитик'
    >>> csv_reader('unittest.csv')[0].vacancies_objects[0].area_name
    'Караганда'
    >>> type(csv_reader('unittest.csv')[0].vacancies_objects[0].published_at).__name__
    'datetime'
    >>> csv_reader('unittest.csv')[0].vacancies_objects[0].published_at.strftime("%Y-%m-%dT%H:%M:%S%z")
    '2015-07-13T17:34:36+0300'
    >>> type(csv_reader('unittest.csv')[0].vacancies_objects[0].salary).__name__
    'Salary'
    >>> csv_reader('unittest.csv')[0].vacancies_objects[0].salary.salary_from
    1230
    >>> csv_reader('unittest.csv')[0].vacancies_objects[0].salary.salary_to
    9870
    >>> csv_reader('unittest.csv')[0].vacancies_objects[0].salary.salary_currency
    'KGS'
    """

    with open(file_name, encoding="utf-8-sig") as file:
        reader = list(csv.reader(file))
        try:
            list_naming = reader.pop(0)
        except Exception:
            list_naming = None
        data_set = DataSet(file_name, reader)
        file.close()
        return data_set, list_naming


def main():
    name_file = input('Введите название файла: ')
    profession_name = input('Введите название профессии: ')
    data_set, list_naming = csv_reader(name_file)
    data_set = DataSet('unittest.csv', [['IT аналитик', '32000.0', '56000.0', 'RUR', 'Санкт-Петербург', '2007-12-03T17:34:36+0300'],
                ['PHP-программист', '41000.0', '67000.0', 'RUR', 'Москва', '2007-12-03T22:39:07+0300'],
                ['Web-программист', '30000.0', '40000.0', 'RUR', 'Москва', '2008-12-03T19:10:20+0300']])
    if list_naming is None:
        print('Пустой файл')
    elif len(data_set.vacancies_objects) == 0:
        print('Нет данных')
    else:
        statistic = data_set.calculate_statistics(profession_name)
        sheet1_headlines = {'A': 'Год', 'B': 'Средняя зарплата', 'C': f'Средняя зарплата - {profession_name}',
                            'D': 'Количество вакансий', 'E': f'Количество вакансий - {profession_name}'}
        sheet2_headlines = {'A': 'Город', 'B': 'Уровень зарплат', 'D': 'Город', 'E': 'Доля вакансий'}
        report = Report(['Статистика по годам', 'Статистика по городам'], [sheet1_headlines, sheet2_headlines])
        report.generate_excel(statistic)


def get_tabular_statistics(name_file, profession_name, sheet_titles, sheet_headlines):
    """Метод запускающий программу.

    Args:
       name_file (str): Название файла
       profession_name (str): Название профессии
       sheet_titles (list[str]): Названия листов таблицы
       sheet_headlines (list[dict[str: str]]): Заголовки, которые присваиваются определенным столбцам в первой строчке
    """
    data_set, list_naming = csv_reader(name_file)
    if list_naming is None:
        print('Пустой файл')
    elif len(data_set.vacancies_objects) == 0:
        print('Нет данных')
    else:
        statistic = data_set.calculate_statistics(data_set, profession_name)
        report = Report(sheet_titles, sheet_headlines)
        report.generate_excel(statistic)


if __name__ == '__main__':
    # main()
    doctest.testmod()
